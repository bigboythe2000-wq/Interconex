import pandas as pd
import os
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

def formato_texto(valor):
    """Añade comilla simple para forzar formato texto en Excel."""
    if valor is None:
        return ""
    return f"'{str(valor)}"

def guardar_en_excel(nuevo_df, nombre_archivo, nombre_hoja):
    # ... (el resto de la función guardar_en_excel se mantiene igual)
    if os.path.exists(nombre_archivo):
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            archivo_existente = pd.read_excel(nombre_archivo, sheet_name=None, dtype=str)
            if nombre_hoja in archivo_existente:
                df_hoja = pd.concat([archivo_existente[nombre_hoja], nuevo_df], ignore_index=True)
            else:
                df_hoja = nuevo_df
            df_hoja.to_excel(writer, sheet_name=nombre_hoja, index=False)
            aplicar_formato(writer.sheets[nombre_hoja])
    else:
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            nuevo_df.to_excel(writer, sheet_name=nombre_hoja, index=False)
            aplicar_formato(writer.sheets[nombre_hoja])

def aplicar_formato(workSheet):
    # ... (función aplicar_formato se mantiene igual)
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    for column_cells in workSheet.columns:
        longitud_maxima = 0
        for cell in column_cells:
            if cell.value is not None:
                longitud_maxima = max(longitud_maxima, len(str(cell.value)))
            cell.border = borde_fino
        letra_columna = column_cells[0].column_letter
        workSheet.column_dimensions[letra_columna].width = longitud_maxima + 2

def escribirTransferencia(Transferencia, fechaHora, estadoTransferencia, referencia, descripcion, nombreArchivo):
    df = pd.DataFrame([{
        'Cuenta Pagadora': formato_texto(Transferencia.cuentaEmisor),
        'ID Pagador': formato_texto(Transferencia.idEmisor),
        'Cuenta Receptora': formato_texto(Transferencia.cuentaReceptor),
        'Teléfono Receptor': formato_texto(Transferencia.telefonoReceptor),
        'ID Receptor': formato_texto(Transferencia.idReceptor),
        'Canal': Transferencia.canal,
        'ID Consumidor': formato_texto(Transferencia.idConsumidor),
        'Monto': Transferencia.monto, # El monto puede dejarse como número si deseas hacer sumas
        'Moneda': Transferencia.moneda,
        'Beneficiario': Transferencia.beneficiario,
        'Concepto': Transferencia.concepto,
        'Estado': estadoTransferencia,
        'Fecha y Hora': fechaHora,
        'Referencia': formato_texto(referencia),
        'Descripcion': descripcion
    }])
    guardar_en_excel(df, nombreArchivo, 'Transferencias')

def escribirPagoMovil(pagoMovil, fechaHora, estadoTransferencia, referencia, nombreArchivo):
    df = pd.DataFrame([{
        'Cuenta Pagadora': formato_texto(pagoMovil.cuentaOrdenante),
        'ID Pagador': formato_texto(pagoMovil.idOrdenante),
        'Teléfono Pagador': formato_texto(pagoMovil.telefonoOrdenante),
        'Teléfono Receptor': formato_texto(pagoMovil.telefonoReceptor),
        'ID Receptor': formato_texto(pagoMovil.idReceptor),
        'ID Consumidor': formato_texto(pagoMovil.idConsumidor),
        'Canal': pagoMovil.canal,
        'Monto': pagoMovil.monto,
        'Moneda': pagoMovil.moneda,
        'Estado': estadoTransferencia,
        'Fecha and Hora': fechaHora,
        'Referencia': formato_texto(referencia)
    }])
    guardar_en_excel(df, nombreArchivo, 'Pago Móvil')

def reiniciar_archivo_solicitudes(nombre_archivo="Solicitudes.xlsx"):
    """Elimina el archivo si existe, lo crea vacío con sus columnas

    y le aplica un estilo profesional a los encabezados.
    """

    # 1. Validación y reinicio del archivo
    if os.path.exists(nombre_archivo):
        try:
            os.remove(nombre_archivo)
            print(f"Archivo previo '{nombre_archivo}' eliminado.")
        except PermissionError:
            print(
                f"⚠️ Error: Cierra el archivo '{nombre_archivo}' en Excel antes de ejecutar."
            )
            return

    # 2. Definición de la estructura (columnas vacías)
    hojas = {
        "Transferencias": pd.DataFrame(
            columns=[
                "Id Emisor",
                "Cuenta Emisor",
                "Id Receptor",
                "Cuenta Receptor",
                "Teléfono Receptor",
                "banco",
                "monto",
                "moneda",
                "canal",
                "id Consumidor",
                "Beneficiario",
                "Concepto",
            ]
        ),
        "Pago Móvil": pd.DataFrame(
            columns=[
                "Id Ordenante",
                "Cuenta Ordenante",
                "Teléfono Ordenante",
                "Id Receptor",
                "Teléfono Receptor",
                "Banco (código)",
                "Banco (nombre)",
                "monto",
                "moneda",
                "canal",
                "id Consumidor",
                "Concepto",
            ]
        ),
        "Consulta Transferencias": pd.DataFrame(
            columns=[
                "id_Cliente",
                "Id_Canal",
                "id_Usuario",
                "Id_Consumidor",
                "fecha",
                "documentoContraparte",
                "referencia",
                "cuenta",
                "bancoContraparte",
                "monto",
                "telefonoContraparte",
                "tipoConsulta",
                "otroBanco",
                "posicionInicial",
            ]
        ),
        "Consulta Límites Pago Móvil": pd.DataFrame(
            columns=[
                "id_Cliente",
                "id_Canal",
                "id_Consumidor",
                "telefono_Cliente",
            ]
        ),
        "Consulta Limites Transferencia": pd.DataFrame(
            columns=["id_Cliente", "id_Canal", "id_Consumidor"]
        ),
        "Lista De Movimientos": pd.DataFrame(
            columns=[
                "rifcliente",
                "idcanal",
                "rifconsumidor",
                "cuenta",
                "fecha",
                "posicion",
                "tipolistado",
            ]
        ),
    }

    # 3. Estilos visuales (Paleta Azul Corporativo)
    fuente_encabezado = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    relleno_encabezado = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )  # Azul Oscuro
    alineacion_encabezado = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # Borde delgado gris para las celdas de títulos
    borde_delgado = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="medium", color="1F497D"),  # Borde inferior un pelo más grueso
    )

    # 4. Escritura y Estilizado en Excel
    with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

            # Accedemos a la hoja de openpyxl para aplicar los estilos de celda
            ws = writer.sheets[nombre_hoja]

            # Ajustar alto de la fila del encabezado para que respire
            ws.row_dimensions[1].height = 28

            # Aplicar estilos a cada celda de la primera fila (los títulos)
            for col_num, column_title in enumerate(df.columns, 1):
                celda = ws.cell(row=1, column=col_num)
                celda.font = fuente_encabezado
                celda.fill = relleno_encabezado
                celda.alignment = alineacion_encabezado
                celda.border = borde_delgado

                # Autoajustar el ancho de la columna según el largo del texto del título
                largo_texto = len(str(column_title))
                col_letter = get_column_letter(col_num)
                # Le damos un margen de +4 caracteres para que no quede pegado
                ws.column_dimensions[col_letter].width = max(
                    largo_texto + 4, 15
                )

    print(
        f" ¡Listo! Archivo '{nombre_archivo}' reiniciado con éxito."
    )

def guardar_payload_en_excel(payload, nombre_archivo, nombre_hoja):
    """
    Recibe un diccionario (payload), lo convierte a DataFrame 
    y lo añade a la hoja indicada.
    """
    # Convertimos el diccionario a una lista de un solo elemento para crear el DataFrame
    # Usamos [payload] para que pandas entienda que es una fila
    df = pd.DataFrame([payload])
    
    # Reutilizamos la función anterior
    guardar_en_excel(df, nombre_archivo, nombre_hoja)
    print(f"✅ Payload guardado exitosamente en '{nombre_hoja}'.")

# guardar_payload_en_excel({    
#   "idCliente": "J1234567",
#   "idUsuario": "usuario_beca",
#   "idTerminal": "terminal_beca",
#   "idCanal": "01",
#   "idConsumidor": "J1234567",
#   "ipOrigen": "111.111.11.1",
#   "idPagador": "V1234567",
#   "telfPagador": "584141234567",
#   "telfReceptor": "584141234567",
#   "ctaReceptor": "0115xxxxxxxxxxxxxxxx",
#   "moneda": "VES",
#   "serialOperacion": "91005224827",
#   "monto": "10.5",
#   "codBanco": "0115"
# },"Resultados.xlsx", "prueba")